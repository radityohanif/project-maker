from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from quote_maker.core.models import QuoteSpec
from shared.utils.files import ensure_parent


def _number_format(currency: str) -> str:
    safe = currency.replace('"', "")
    return f'"{safe}" #,##0.00'


def _percent_format() -> str:
    return "0.0%"


def render(spec: QuoteSpec, path: Path) -> Path:
    """Write a quotation workbook with formulas for amounts, subtotals, and summary."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F2937")
    section_fill = PatternFill(fill_type="solid", fgColor="E5E7EB")
    total_fill = PatternFill(fill_type="solid", fgColor="FEF3C7")
    grand_fill = PatternFill(fill_type="solid", fgColor="FCA5A5")
    num_fmt = _number_format(spec.currency)
    pct_fmt = _percent_format()

    ws["A1"] = "Quotation"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Project: {spec.meta.name}"
    if spec.meta.client:
        ws["A3"] = f"Client: {spec.meta.client}"
    if spec.meta.date:
        ws["A4"] = f"Date: {spec.meta.date}"

    header_row = 6
    headers = ["Position", "Cost", "Qty", "Contract", "Amount"]
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    r = header_row + 1
    section_subtotal_rows: list[int] = []

    for section in spec.sections:
        sc = ws.cell(r, 1, section.title)
        sc.font = Font(bold=True)
        sc.fill = section_fill
        for col in range(1, 6):
            cell = ws.cell(r, col)
            cell.fill = section_fill
            cell.border = border
        r += 1

        first_item_row = r
        for item in section.items:
            ws.cell(r, 1, item.position)
            cost_cell = ws.cell(r, 2, item.unit_cost)
            cost_cell.number_format = num_fmt
            qty_cell = ws.cell(r, 3, item.qty)
            qty_cell.alignment = Alignment(horizontal="center")
            contract_cell = ws.cell(r, 4, item.contract)
            contract_cell.alignment = Alignment(horizontal="center")
            amount_cell = ws.cell(r, 5, f"=B{r}*C{r}*D{r}")
            amount_cell.number_format = num_fmt
            for col in range(1, 6):
                ws.cell(r, col).border = border
            r += 1

        last_item_row = r - 1
        ws.cell(r, 1, f"Subtotal — {section.title}").font = Font(italic=True)
        sub_amount = ws.cell(
            r,
            5,
            f"=SUM(E{first_item_row}:E{last_item_row})" if last_item_row >= first_item_row else "=0",
        )
        sub_amount.number_format = num_fmt
        sub_amount.font = Font(bold=True)
        for col in range(1, 6):
            ws.cell(r, col).border = border
        section_subtotal_rows.append(r)
        r += 1

    r += 1
    totals_header = ws.cell(r, 1, "Item")
    totals_header.font = header_font
    totals_header.fill = header_fill
    totals_header.border = border
    value_header = ws.cell(r, 2, "Value")
    value_header.font = header_font
    value_header.fill = header_fill
    value_header.border = border
    rate_header = ws.cell(r, 3, "Rate")
    rate_header.font = header_font
    rate_header.fill = header_fill
    rate_header.border = border
    rate_header.alignment = Alignment(horizontal="center", vertical="center")
    r += 1

    if not section_subtotal_rows:
        subtotal_expr = "=0"
    elif len(section_subtotal_rows) == 1:
        subtotal_expr = f"=E{section_subtotal_rows[0]}"
    else:
        subtotal_expr = "=" + "+".join(f"E{row}" for row in section_subtotal_rows)

    subtotal_summary_row = r
    ws.cell(r, 1, "Subtotal")
    ws.cell(r, 2, subtotal_expr)
    ws.cell(r, 2).number_format = num_fmt
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = total_fill
        c.border = border
    r += 1

    risk_row = r
    ws.cell(r, 1, "Risk")
    ws.cell(r, 2, f"=B{subtotal_summary_row}*C{risk_row}")
    ws.cell(r, 2).number_format = num_fmt
    risk_rate_cell = ws.cell(r, 3, spec.risk)
    risk_rate_cell.number_format = pct_fmt
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = total_fill
        c.border = border
    r += 1

    markup_row = r
    ws.cell(r, 1, "Markup")
    ws.cell(r, 2, f"=(B{subtotal_summary_row}+B{risk_row})*C{markup_row}")
    ws.cell(r, 2).number_format = num_fmt
    margin_rate_cell = ws.cell(r, 3, spec.markup)
    margin_rate_cell.number_format = pct_fmt
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = total_fill
        c.border = border
    r += 1

    pretax_row = r
    ws.cell(r, 1, "Pre-tax")
    ws.cell(r, 2, f"=B{subtotal_summary_row}+B{risk_row}+B{markup_row}")
    ws.cell(r, 2).number_format = num_fmt
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = total_fill
        c.border = border
    r += 1

    tax_row = r
    ws.cell(r, 1, f"Tax ({spec.tax * 100:.1f}%)")
    ws.cell(r, 2, f"=B{pretax_row}*C{tax_row}")
    ws.cell(r, 2).number_format = num_fmt
    tax_rate_cell = ws.cell(r, 3, spec.tax)
    tax_rate_cell.number_format = pct_fmt
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = total_fill
        c.border = border
    r += 1

    grand_row = r
    ws.cell(r, 1, "Grand Total")
    ws.cell(r, 2, f"=B{pretax_row}+B{tax_row}")
    ws.cell(r, 2).number_format = num_fmt
    grand_font = Font(bold=True)
    for col in (1, 2, 3):
        c = ws.cell(r, col)
        c.fill = grand_fill
        c.border = border
        c.font = grand_font

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 20

    ensure_parent(path)
    wb.save(path)
    return path
