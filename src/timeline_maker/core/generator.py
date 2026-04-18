from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from shared.utils.files import ensure_parent
from timeline_maker.core.models import PhaseRow, TaskRow, TimelineSpec


def col_for_week(month_idx: int, week_idx: int, weeks_per_month: int) -> int:
    """Map a ``(month, week)`` slot to its 1-based Excel column.

    Columns 1 and 2 are reserved for ``No`` and ``Activity``; week columns start at 3.
    """
    return 3 + month_idx * weeks_per_month + week_idx


def build_workbook(spec: TimelineSpec, path: Path) -> Path:
    """Write a single-sheet Gantt-style workbook from ``spec`` to ``path``.

    Returns the output path for convenience.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = spec.meta.sheet_title[:31]

    num_months = spec.timeline.num_months
    wpm = spec.timeline.weeks_per_month
    freeze_months = frozenset(spec.timeline.freeze_month_indices)

    thin = Side(style="thin", color="808080")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, size=11)
    phase_fill = PatternFill(fill_type="solid", fgColor="D9D9D9")
    active_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    freeze_header_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    freeze_col_fill = PatternFill(fill_type="solid", fgColor="EDEDED")
    meta_font = Font(size=10)

    ws["A1"], ws["B1"] = "Version:", spec.meta.version
    ws["A2"], ws["B2"] = "Last updated:", spec.meta.updated
    ws["A3"], ws["B3"] = "Calendar:", spec.meta.calendar_note or _default_calendar_note(spec)
    ws["A4"], ws["B4"] = "Freeze months:", spec.meta.freeze_note or _default_freeze_note(spec)
    for row in range(1, 5):
        for col in range(1, 4):
            ws.cell(row, col).font = meta_font

    header_row = 6
    week_row = 7
    last_col = 2 + num_months * wpm

    ws.cell(header_row, 1, "No")
    ws.cell(header_row, 2, "Activity")
    for m in range(num_months):
        c0 = col_for_week(m, 0, wpm)
        c1 = col_for_week(m, wpm - 1, wpm)
        ws.merge_cells(start_row=header_row, start_column=c0, end_row=header_row, end_column=c1)
        cell = ws.cell(header_row, c0, spec.months[m])
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if m in freeze_months:
            cell.fill = freeze_header_fill
        for w in range(wpm):
            cc = col_for_week(m, w, wpm)
            wcell = ws.cell(week_row, cc, w + 1)
            wcell.font = Font(size=9)
            wcell.alignment = Alignment(horizontal="center")
            if m in freeze_months:
                wcell.fill = freeze_header_fill

    for c in range(1, last_col + 1):
        ws.cell(header_row, c).border = border
        ws.cell(week_row, c).border = border

    data_start = 8
    no = 0
    r = data_start
    for item in spec.rows:
        if isinstance(item, PhaseRow):
            ws.cell(r, 1, "")
            ws.cell(r, 2, item.label)
            ws.cell(r, 2).font = Font(bold=True)
            ws.cell(r, 2).fill = phase_fill
            for m in range(num_months):
                for w in range(wpm):
                    cc = col_for_week(m, w, wpm)
                    fill = freeze_col_fill if m in freeze_months else phase_fill
                    ws.cell(r, cc, "").fill = fill
            r += 1
            continue

        assert isinstance(item, TaskRow)
        no += 1
        slot_set = spec.effective_task_slots(item)
        ws.cell(r, 1, no)
        ws.cell(r, 2, item.label)
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for m in range(num_months):
            for w in range(wpm):
                cc = col_for_week(m, w, wpm)
                cell = ws.cell(r, cc, "")
                if m in freeze_months:
                    cell.fill = freeze_col_fill
                elif (m, w) in slot_set:
                    cell.fill = active_fill
        r += 1

    last_row = r - 1
    for row in range(data_start, last_row + 1):
        for col in range(1, last_col + 1):
            ws.cell(row, col).border = border

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 68
    for m in range(num_months):
        for w in range(wpm):
            cc = get_column_letter(col_for_week(m, w, wpm))
            ws.column_dimensions[cc].width = 4.0

    ws.freeze_panes = f"C{data_start}"
    ensure_parent(path)
    wb.save(path)
    return path


def _default_calendar_note(spec: TimelineSpec) -> str:
    t = spec.timeline
    return (
        f"{t.start_year}-{t.start_month:02d} start, {t.num_months} month column(s), "
        f"{t.weeks_per_month} symbolic week slot(s) per month (not ISO weeks)."
    )


def _default_freeze_note(spec: TimelineSpec) -> str:
    if not spec.timeline.freeze_month_indices:
        return "None (all month columns are active for work slots)."
    return f"No work shading on month indices: {sorted(spec.timeline.freeze_month_indices)}."
