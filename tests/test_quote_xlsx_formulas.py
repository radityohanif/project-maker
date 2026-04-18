from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from quote_maker.core.models import QuoteItem, QuoteSection, QuoteSpec
from quote_maker.core.renderer import render
from shared.schemas.common import ProjectMeta


def _tiny_spec() -> QuoteSpec:
    return QuoteSpec(
        meta=ProjectMeta(name="P", client="C", date="2026-01-01"),
        currency="IDR",
        markup=0.10,
        risk=0.05,
        tax=0.11,
        sections=[
            QuoteSection(
                title="Sec",
                items=[QuoteItem(position="Role", qty=2, unit_cost=100, contract=3)],
            ),
        ],
    )


def test_quotation_workbook_uses_excel_formulas(tmp_path: Path) -> None:
    path = tmp_path / "quotation.xlsx"
    render(_tiny_spec(), path)

    wb = load_workbook(path, data_only=False)
    ws = wb.active

    # Header row 6; section title 7; first line item row 8
    assert ws["E8"].data_type == "f"
    assert ws["E8"].value == "=B8*C8*D8"

    assert ws["E9"].data_type == "f"
    assert "SUM(E8:E8)" in (ws["E9"].value or "")

    # Summary block: locate "Subtotal" label in column A
    sub_row = None
    for row in range(1, 30):
        v = ws.cell(row, 1).value
        if v == "Subtotal":
            sub_row = row
            break
    assert sub_row is not None
    assert ws.cell(sub_row, 2).data_type == "f"
    assert "E9" in (ws.cell(sub_row, 2).value or "")

    risk_row = sub_row + 1
    assert ws.cell(risk_row, 1).value == "Risk"
    assert ws.cell(risk_row, 2).data_type == "f"
    assert "B" in (ws.cell(risk_row, 2).value or "")

    markup_row = risk_row + 1
    assert ws.cell(markup_row, 1).value == "Markup"
    assert ws.cell(markup_row, 2).data_type == "f"
    val = ws.cell(markup_row, 2).value or ""
    assert "(" in val and "*" in val

    pretax_row = markup_row + 1
    assert ws.cell(pretax_row, 1).value == "Pre-tax"
    assert ws.cell(pretax_row, 2).data_type == "f"

    tax_row = pretax_row + 1
    assert ws.cell(tax_row, 2).data_type == "f"

    grand_row = tax_row + 1
    assert ws.cell(grand_row, 1).value == "Grand Total"
    assert ws.cell(grand_row, 2).data_type == "f"
