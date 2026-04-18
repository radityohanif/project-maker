from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from timeline_maker.core.generator import build_workbook
from timeline_maker.core.parser import parse_file

EXAMPLE = Path(__file__).resolve().parents[1] / "examples" / "timeline.yaml"


def test_timeline_generate_produces_openable_xlsx(tmp_path: Path) -> None:
    spec = parse_file(EXAMPLE)
    out = tmp_path / "timeline.xlsx"
    build_workbook(spec, out)
    assert out.exists()
    wb = load_workbook(out)
    ws = wb.active
    assert ws.cell(6, 1).value == "No"
    assert ws.cell(6, 2).value == "Activity"
